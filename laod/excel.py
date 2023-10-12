import re
from enum import Enum

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from engine.data import PageKind, TablePage, EnumPage, ConfigPage, ValueKind
from engine.load import LoadFile


def get_cell(cell: Cell):
    value = cell.value
    if cell.data_type != "s":
        return value
    if value.startswith("#"):
        return None
    return value


# 获取页签类型
def get_page_kind(page_kind: str):
    for kind_enum in PageKind:
        if kind_enum.name == page_kind:
            return kind_enum
    return PageKind.Null


class ExcelHead:
    class Kind(Enum):
        Default = 0
        Struct = 1
        Array = 2
        Index = 3

    idx: int
    key: str
    desc: str
    link: str
    kind: Kind


def get_key(cell: Cell):
    return get_cell(cell)


def get_heads(ws: Worksheet):
    heads: list[ExcelHead] = list()
    key_row = next(ws.iter_rows(2))
    desc_row = next(ws.iter_rows(3))
    for cell in key_row:
        head = ExcelHead()
        head.idx = cell.col_idx
        head.key = get_key(cell)
        head.desc = get_cell(desc_row[head.idx - 1])
        head.kind = ExcelHead.Kind.Default
        heads.append(head)
    return heads


def to_table(page_name: str, ws: Worksheet):
    table_page: TablePage = TablePage(page_name)
    table_heads = get_heads(ws)
    return table_page


def to_enum(page_name: str, ws: Worksheet):
    enum_page: EnumPage = EnumPage(page_name)
    return enum_page


def to_config(page_name: str, ws: Worksheet):
    config_page: ConfigPage = ConfigPage(page_name)
    return config_page


def to_page(ws: Worksheet) -> any:
    page_tag = get_cell(ws.cell(1, 1))
    if page_tag is None:
        return None
    tag_match = re.match(r"(.*)<(.*)>", page_tag)
    if tag_match and len(tag_match.groups()) != 2:
        return None
    page_kind = get_page_kind(tag_match.group(1))
    page_name = tag_match.group(2)
    if page_kind == PageKind.Table:
        return to_table(page_name, ws)
    elif page_kind == PageKind.Enum:
        return to_enum(page_name, ws)
    elif page_kind == PageKind.Config:
        return to_config(page_name, ws)
    return None


class ExcelLoad(LoadFile):

    def on_load(self) -> []:
        wb: Workbook = openpyxl.load_workbook(self.path)
        for ws in wb:
            page = to_page(ws)
            if page is not None:
                print(page)
                yield page
        wb.close()
