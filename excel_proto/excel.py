import re

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from engine.data import Table, Config, Enumerate
from engine.data.check import Check
from engine.data.enumerate import EnumerateRow
from engine.data.table import TableRow
from engine.load import LoadFile
from engine.data.base import string_to_data_kind, DataKind, Head, Value, data_type_to_value_kind


def get_cell(cell: Cell):
    value = cell.value
    if value is None:
        return None
    if cell.data_type != "s":
        return value
    if value.startswith("#"):
        return None
    return value


def get_head_dict(ws: Worksheet):
    # Key
    key_dict: dict[int, str] = dict()
    key_row = next(ws.iter_rows(2))
    for cell in key_row:
        cell_str = get_cell(cell)
        if cell_str:
            key_dict[cell.col_idx] = cell_str

    # Desc
    desc_dict: dict[int, str] = dict()
    desc_row = next(ws.iter_rows(3))
    for cell in desc_row:
        if cell.col_idx in key_dict.keys():
            desc_dict[cell.col_idx] = get_cell(cell)

    # Head
    head_dict: dict[int, Head] = dict()
    for idx in key_dict.keys():
        head = Head(idx)
        head.set_key(key_dict.get(idx))
        head.set_desc(desc_dict.get(idx))
        head_dict[idx] = head
    return head_dict


def to_data(ws: Worksheet):
    info_cell: Cell = ws.cell(1, 1)
    cell_value = get_cell(info_cell)
    if cell_value is None:
        return None
    match = re.match(r"(.*)<(.*)>", cell_value)
    if match and len(match.groups()) != 2:
        return None
    kind = string_to_data_kind(match.group(1))
    name = match.group(2)
    if kind == DataKind.Table:
        return to_table(name, ws)
    elif kind == DataKind.Config:
        return to_config(name, ws)
    elif kind == DataKind.Enum:
        return to_enumerate(name, ws)
    elif kind == DataKind.Check:
        return to_check(name, ws)
    return None


def to_table(name: str, ws: Worksheet):
    table = Table(name)

    # Head
    head_dict = get_head_dict(ws)

    # Values
    row_list: list[TableRow] = list()
    for value_row in ws.iter_rows(4):
        table_row = TableRow()
        row_list.append(table_row)
        for cell in value_row:
            if cell.col_idx in head_dict.keys():
                value = Value(head_dict[cell.col_idx])
                value.set_kind(data_type_to_value_kind(cell.data_type))
                value.set_meta(get_cell(cell))
                table_row.set_value(value)

    table.set_rows(row_list)
    return table


def to_config(name: str, ws: Worksheet):
    config = Config(name)

    # Head
    head_dict = get_head_dict(ws)

    # Key
    key_idx_dict: dict[str: int] = dict()
    for head in head_dict.values():
        key_idx_dict[head.key] = head.idx - 1

    # Values
    elements: dict[str, Value] = dict()
    row_idx = 4
    for value_row in ws.iter_rows(row_idx):
        config_head: Head = Head(row_idx)
        value = Value(config_head)
        key_idx = key_idx_dict.get("Key")
        if key_idx is None:
            continue
        value_key = get_cell(value_row[key_idx])
        if value_key is None:
            continue
        value_key = str(value_key)
        config_head.set_key(value_key)
        desc_idx = key_idx_dict.get("Desc")
        if desc_idx:
            value_desc = get_cell(value_row[desc_idx])
            config_head.set_desc(str(value_desc))
        value_idx = key_idx_dict.get("Value")
        if value_idx:
            value_cell = value_row[value_idx]
            value.set_meta(get_cell(value_cell))
            value.set_kind(data_type_to_value_kind(value_cell.data_type))
        elements[value_key] = value
        row_idx += 1

    config.set_elements(elements)
    return config


def to_enumerate(name: str, ws: Worksheet):
    enum = Enumerate(name)

    # Head
    head_dict = get_head_dict(ws)

    # Key
    key_idx_dict: dict[str: int] = dict()
    for head in head_dict.values():
        key_idx_dict[head.key] = head.idx - 1

    # Values
    elements: dict[str, EnumerateRow] = dict()
    for value_row in ws.iter_rows(4):
        key_idx = key_idx_dict.get("Key")
        if key_idx is None:
            continue
        value_key = get_cell(value_row[key_idx])
        if value_key is None:
            continue
        value_key = str(value_key)
        element = EnumerateRow(value_key)
        desc_idx = key_idx_dict.get("Desc")
        if desc_idx:
            element.set_desc(str(get_cell(value_row[desc_idx])))
        value_idx = key_idx_dict.get("Value")
        if value_idx:
            element.set_value(int(get_cell(value_row[value_idx])))
        text_idx = key_idx_dict.get("Text")
        if text_idx:
            element.set_text(str(get_cell(value_row[text_idx])))
        elements[value_key] = element

    enum.set_elements(elements)
    return enum


def to_check(name: str, ws: Worksheet):
    check = Check(name)
    return check


class ExcelFile(LoadFile):
    def __init__(self, path: str):
        super().__init__(path)

    def on_load(self) -> []:
        print(self.path)
        wb: Workbook = openpyxl.load_workbook(self.path)
        for ws in wb:
            data = to_data(ws)
            if data is not None:
                print(data)
                yield data
        wb.close()
